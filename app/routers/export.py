import io
from datetime import datetime

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models.dishonour import Dishonour
from app.models.payer import Payer
from app.models.payment import Payment

router = APIRouter()

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1E293B")  # slate-800


def _style_header_row(ws, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


@router.get("/export/dishonours")
async def export_dishonours(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Dishonour, Payer, Payment)
        .outerjoin(Payer, Dishonour.payer_id == Payer.id)
        .outerjoin(Payment, Dishonour.payment_id == Payment.id)
        .order_by(Dishonour.created_at.desc())
    )
    rows = result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Failed Payments"

    headers = [
        "ID", "Payer Name", "Email", "Amount (AUD)", "Reason Code",
        "Reason", "Action Taken", "Status", "Recovery Probability",
        "AI Explanation", "Date",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for dishonour, payer, payment in rows:
        ws.append([
            dishonour.id,
            payer.name if payer else "Unknown",
            payer.email if payer else "",
            f"${(payment.amount_cents if payment else 0) / 100:.2f}",
            dishonour.reason_code,
            dishonour.reason_label,
            dishonour.action_taken or "",
            dishonour.status,
            "",
            dishonour.claude_explanation or "",
            dishonour.created_at.strftime("%d/%m/%Y %H:%M") if dishonour.created_at else "",
        ])

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"retryly-dishonours-{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/payers")
async def export_payers(db: AsyncSession = Depends(get_db)):
    payer_result = await db.execute(select(Payer).order_by(Payer.created_at.desc()))
    payers = payer_result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Payers"

    headers = [
        "ID", "Name", "Email", "Phone",
        "Payments On Time", "Payments Failed", "Recovery Rate (%)", "Risk Score",
        "Created At",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for payer in payers:
        history = payer.payment_history or {}
        on_time = history.get("on_time", 0)
        failures = history.get("failures", 0)
        total = on_time + failures
        rate = round(on_time / total * 100, 1) if total > 0 else 100.0
        if failures >= 2:
            risk = "High"
        elif failures == 1:
            risk = "Medium"
        else:
            risk = "Low"
        ws.append([
            payer.id,
            payer.name,
            payer.email,
            payer.phone or "",
            on_time,
            failures,
            rate,
            risk,
            payer.created_at.strftime("%d/%m/%Y") if payer.created_at else "",
        ])

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"retryly-payers-{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
